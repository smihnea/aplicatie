"""Extract actual URLs from Excel hyperlink formulas."""

import re
import openpyxl
from typing import List, Tuple, Optional


class ExcelHyperlinkExtractor:
    """Extract hyperlinks from Excel files, handling formulas and direct links."""
    
    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path
        
    def extract_hyperlinks_from_sheet(self, sheet_name: str, column_name: str = "FISA TEHNICA") -> List[Tuple[int, str, str]]:
        """
        Extract hyperlinks from a specific sheet and column.
        
        Args:
            sheet_name: Name of the Excel sheet
            column_name: Name of the column containing hyperlinks
            
        Returns:
            List of (row_index, display_text, actual_url) tuples
        """
        hyperlinks = []
        
        try:
            wb = openpyxl.load_workbook(self.excel_file_path, data_only=False)  # Keep formulas
            
            if sheet_name not in wb.sheetnames:
                return hyperlinks
                
            sheet = wb[sheet_name]
            
            # Find the column index for the hyperlink column
            col_index = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header and column_name.upper() in str(header).upper():
                    col_index = col
                    break
                    
            if col_index is None:
                return hyperlinks
                
            # Extract hyperlinks from the column
            for row in range(2, sheet.max_row + 1):  # Skip header row
                cell = sheet.cell(row=row, column=col_index)
                
                if cell.value:
                    cell_value = str(cell.value)
                    
                    # Check if it's a HYPERLINK formula
                    if cell_value.startswith('=HYPERLINK'):
                        url = self._extract_url_from_hyperlink_formula(cell_value, sheet, row, col_index)
                        if url:
                            hyperlinks.append((row - 2, cell_value, url))  # row - 2 for 0-based index
                    
                    # Check if cell has direct hyperlink
                    elif cell.hyperlink:
                        hyperlinks.append((row - 2, cell_value, cell.hyperlink.target))
                        
                    # Check if it's a product code that should become a URL
                    elif self._is_product_code(cell_value):
                        url = f"https://new.abb.com/products/{cell_value}"
                        hyperlinks.append((row - 2, cell_value, url))
                        
            wb.close()
            
        except Exception as e:
            print(f"Error extracting hyperlinks from {sheet_name}: {e}")
            
        return hyperlinks
    
    def _extract_url_from_hyperlink_formula(self, formula: str, sheet, row: int, col_index: int) -> Optional[str]:
        """Extract URL from Excel HYPERLINK formula."""
        
        # Example: =HYPERLINK(_xlfn.CONCAT("https://new.abb.com/products/",C2),C2)
        
        # First, try to extract base URL from the formula
        url_match = re.search(r'"(https?://[^"]+)"', formula)
        if not url_match:
            return None
            
        base_url = url_match.group(1)
        
        # Try to find the cell reference (like C2)
        cell_ref_match = re.search(r',([A-Z]+\d+)\)', formula)
        if cell_ref_match:
            cell_ref = cell_ref_match.group(1)
            
            # Convert cell reference to coordinates
            try:
                ref_col = 0
                for char in cell_ref:
                    if char.isalpha():
                        ref_col = ref_col * 26 + (ord(char.upper()) - ord('A') + 1)
                    else:
                        ref_row = int(cell_ref[cell_ref.index(char):])
                        break
                
                # Get the value from the referenced cell  
                ref_cell = sheet.cell(row=ref_row, column=ref_col)
                if ref_cell.value:
                    # Construct final URL
                    return f"{base_url}{ref_cell.value}"
                    
            except Exception:
                pass
                
        # If we can't parse the reference, try to extract from current row
        # Look for product code in the same row (usually column C)
        code_cell = sheet.cell(row=row, column=3)  # Column C = index 3
        if code_cell.value:
            return f"{base_url}{code_cell.value}"
            
        return base_url  # Return base URL if we can't get the product code
    
    def _is_product_code(self, value: str) -> bool:
        """Check if a value looks like a product code."""
        if not isinstance(value, str) or len(value) < 5:
            return False
            
        # Remove common separators
        clean_value = value.replace('-', '').replace('_', '').replace(' ', '')
        
        # Should be alphanumeric and reasonable length
        return clean_value.isalnum() and 5 <= len(clean_value) <= 50