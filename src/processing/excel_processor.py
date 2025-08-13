"""Core Excel file processing functionality."""

import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

from models.product_data import ExcelData, ExcelMetadata, ProductLink
from models.processing_result import ProcessingResult, ProcessingStatus
from utils.logger import get_logger
from .excel_hyperlink_extractor import ExcelHyperlinkExtractor


class ExcelProcessor:
    """Handles Excel file reading, processing, and writing operations."""
    
    def __init__(self):
        self.logger = get_logger(__name__)
    
    def read_excel_file(self, file_path: str) -> ExcelData:
        """
        Read an Excel file and extract all relevant data.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            ExcelData object containing all sheets and metadata
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file is not a valid Excel file
        """
        self.logger.info(f"Reading Excel file: {file_path}")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            # Get file metadata
            file_stats = os.stat(file_path)
            metadata = ExcelMetadata(
                filename=os.path.basename(file_path),
                file_path=file_path,
                file_size=file_stats.st_size,
                created_at=datetime.fromtimestamp(file_stats.st_ctime),
                modified_at=datetime.fromtimestamp(file_stats.st_mtime)
            )
            
            # Read all sheets from the Excel file
            with pd.ExcelFile(file_path) as excel_file:
                sheet_names = excel_file.sheet_names
                metadata.sheet_names = sheet_names
                
                sheets = {}
                total_rows = 0
                total_columns = 0
                
                for sheet_name in sheet_names:
                    self.logger.debug(f"Reading sheet: {sheet_name}")
                    
                    # Read sheet with proper handling of various data types
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        engine='openpyxl',
                        dtype=str,  # Read everything as string initially to preserve data
                        na_values=[''],
                        keep_default_na=False
                    )
                    
                    sheets[sheet_name] = df
                    total_rows += len(df)
                    total_columns = max(total_columns, len(df.columns))
                
                metadata.total_rows = total_rows
                metadata.total_columns = total_columns
            
            # Create ExcelData object
            excel_data = ExcelData(
                metadata=metadata,
                sheets=sheets,
                original_file_path=file_path
            )
            
            self.logger.info(f"Successfully read Excel file with {len(sheet_names)} sheets, "
                           f"{total_rows} total rows")
            
            return excel_data
            
        except Exception as e:
            self.logger.error(f"Failed to read Excel file {file_path}: {str(e)}")
            raise ValueError(f"Invalid Excel file or read error: {str(e)}")
    
    def extract_links(self, excel_data: ExcelData) -> List[ProductLink]:
        """
        Extract product links from Excel data using hyperlink extractor.
        
        Args:
            excel_data: ExcelData object containing sheet data
            
        Returns:
            List of ProductLink objects found in the Excel sheets
        """
        self.logger.info("Extracting product links from Excel data using hyperlink extractor")
        
        links = []
        skipped_entries = []
        total_items_processed = 0
        
        # Use hyperlink extractor to get real URLs
        extractor = ExcelHyperlinkExtractor(excel_data.original_file_path)
        
        self.logger.info(f"🔗 Starting full extraction from {len(excel_data.sheets)} sheets")
        
        for sheet_index, sheet_name in enumerate(excel_data.sheets.keys(), 1):
            self.logger.info(f"📊 Processing sheet {sheet_index}/{len(excel_data.sheets)}: {sheet_name}")
            
            # Extract hyperlinks from this sheet
            sheet_hyperlinks = extractor.extract_hyperlinks_from_sheet(sheet_name, "FISA TEHNICA")
            self.logger.info(f"  Found {len(sheet_hyperlinks)} hyperlinks in {sheet_name}")
            
            sheet_processed = 0
            for row_index, display_text, actual_url in sheet_hyperlinks:
                total_items_processed += 1
                sheet_processed += 1
                self.logger.debug(f"Found hyperlink: {display_text} -> {actual_url}")
                
                # Validate the extracted URL
                from utils.validators import validate_url
                is_valid, error_msg = validate_url(actual_url)
                
                if is_valid:
                    # Use the validated URL
                    cleaned_url = error_msg if error_msg else actual_url
                    link = ProductLink(
                        url=cleaned_url,
                        row_index=row_index,
                        sheet_name=sheet_name,
                        column_name="FISA TEHNICA"
                    )
                    links.append(link)
                else:
                    # Log skipped entries
                    skipped_entries.append({
                        'value': display_text,
                        'sheet': sheet_name,
                        'row': row_index + 2,  # +2 for Excel row numbering
                        'reason': error_msg,
                        'extracted_url': actual_url
                    })
        
        excel_data.links = links
        excel_data.metadata.links_found = len(links)
        
        # Log summary of processing
        self.logger.info(f"✅ FULL PROCESSING: Processed {total_items_processed} total items from all sheets")
        self.logger.info(f"🔗 Extracted {len(links)} valid product links using hyperlink extractor")
        
        # Log per-sheet summary
        sheet_counts = {}
        for link in links:
            sheet_counts[link.sheet_name] = sheet_counts.get(link.sheet_name, 0) + 1
        
        for sheet_name, count in sheet_counts.items():
            self.logger.info(f"  📊 {sheet_name}: {count} links")
            
        if skipped_entries:
            self.logger.info(f"Skipped {len(skipped_entries)} invalid entries:")
            for entry in skipped_entries[:5]:  # Show first 5 examples
                self.logger.info(f"  Row {entry['row']} in {entry['sheet']}: '{entry['value'][:30]}...' -> '{entry['extracted_url'][:50]}...' - {entry['reason']}")
            if len(skipped_entries) > 5:
                self.logger.info(f"  ... and {len(skipped_entries) - 5} more")
        
        return links
    
    def _find_link_columns(self, df: pd.DataFrame) -> List[str]:
        """Find columns that likely contain web links."""
        link_columns = []
        
        # Common column names that might contain links
        link_indicators = [
            'fisa technica', 'fisa_technica', 'fisatechnica',
            'link', 'url', 'website', 'web', 'href',
            'technical sheet', 'datasheet', 'spec sheet'
        ]
        
        for column in df.columns:
            column_lower = str(column).lower()
            
            # Check if column name matches known link indicators
            if any(indicator in column_lower for indicator in link_indicators):
                link_columns.append(column)
                continue
            
            # Check if column contains URL-like values
            if self._column_contains_urls(df[column]):
                link_columns.append(column)
        
        return link_columns
    
    def _column_contains_urls(self, series: pd.Series) -> bool:
        """Check if a pandas Series contains URL-like values."""
        sample_size = min(10, len(series))
        sample_values = series.dropna().head(sample_size)
        
        url_count = 0
        for value in sample_values:
            if self._is_valid_url(str(value)):
                url_count += 1
        
        # If more than 30% of sample values are URLs, consider it a link column
        return (url_count / len(sample_values)) > 0.3 if len(sample_values) > 0 else False
    
    def _is_valid_url(self, url: str) -> bool:
        """Basic URL validation."""
        url = url.strip()
        if not url:
            return False
        
        # Check for common URL patterns
        url_patterns = [
            url.startswith('http://'),
            url.startswith('https://'),
            url.startswith('www.'),
            ('.' in url and len(url.split('.')[-1]) >= 2)  # Basic domain check
        ]
        
        return any(url_patterns) and len(url) > 5
    
    def write_enhanced_excel(self, excel_data: ExcelData, output_path: str, 
                           preserve_formatting: bool = False) -> str:
        """
        Write enhanced Excel file with extracted product data.
        
        Args:
            excel_data: ExcelData object with extracted information
            output_path: Path where to save the enhanced file
            preserve_formatting: Whether to preserve original formatting
            
        Returns:
            Path to the created enhanced Excel file
        """
        self.logger.info(f"Writing enhanced Excel file: {output_path}")
        
        try:
            # Ensure output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:  # Only create directory if path has a directory component
                os.makedirs(output_dir, exist_ok=True)
            
            if preserve_formatting and self._can_preserve_formatting(excel_data.original_file_path):
                return self._write_with_formatting(excel_data, output_path)
            else:
                return self._write_without_formatting(excel_data, output_path)
                
        except Exception as e:
            self.logger.error(f"Failed to write enhanced Excel file: {str(e)}")
            raise
    
    def _can_preserve_formatting(self, original_path: str) -> bool:
        """Check if we can preserve the original formatting."""
        try:
            # Try to open with openpyxl to see if it's compatible
            openpyxl.load_workbook(original_path)
            return True
        except:
            return False
    
    def _write_with_formatting(self, excel_data: ExcelData, output_path: str) -> str:
        """Write Excel file while preserving original formatting."""
        # Load the original workbook to preserve formatting
        wb = openpyxl.load_workbook(excel_data.original_file_path)
        
        # Define new columns for extracted data
        new_columns = {
            'EAN Code': 'ean',
            'RAL Number': 'ral_number',
            'Net Width (mm)': 'net_width',
            'Net Height (mm)': 'net_height',
            'Net Depth (mm)': 'net_depth',
            'Package Units': 'package_units',
            'Package Weight (kg)': 'package_weight',
            'Data Confidence': 'extraction_confidence',
            'Extraction Method': 'extraction_method'
        }
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            if sheet_name in excel_data.sheets:
                ws = wb[sheet_name]
                
                # Find the rightmost column to add new data
                max_col = ws.max_column
                
                # Add headers for new columns
                header_row = 1  # Assuming first row contains headers
                for i, (header, _) in enumerate(new_columns.items()):
                    ws.cell(row=header_row, column=max_col + i + 1, value=header)
                
                # Add extracted data
                for link in excel_data.get_links_by_sheet(sheet_name):
                    if link.extracted_data:
                        row_num = link.row_index + 2  # +2 because pandas is 0-indexed and we have headers
                        
                        for i, (_, field_name) in enumerate(new_columns.items()):
                            value = getattr(link.extracted_data, field_name, '')
                            if value is not None:
                                ws.cell(row=row_num, column=max_col + i + 1, value=value)
        
        # Save the enhanced workbook
        wb.save(output_path)
        self.logger.info(f"Enhanced Excel file saved with formatting preserved: {output_path}")
        return output_path
    
    def _write_without_formatting(self, excel_data: ExcelData, output_path: str) -> str:
        """Write Excel file with ONLY the extracted data columns (no original data)."""
        self.logger.info(f"📝 Creating results-only Excel file: {output_path}")
        self.logger.info(f"📊 Processing {len(excel_data.sheets)} sheets")
        
        total_links_with_data = 0
        total_links_without_data = 0
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in excel_data.sheets.items():
                self.logger.info(f"🔍 Processing sheet: {sheet_name}")
                
                # Create new dataframe with ONLY the extracted columns (NO original data)
                new_columns = [
                    'Product Code',
                    'EAN Code', 
                    'RAL Number',
                    'Net Width (mm)',
                    'Net Height (mm)', 
                    'Net Depth (mm)',
                    'Package Units',
                    'Package Weight (kg)'
                ]
                
                # Create empty dataframe with new structure
                result_data = []
                
                # Get sheet links and create rows only for products that had links
                sheet_links = excel_data.get_links_by_sheet(sheet_name)
                self.logger.info(f"📋 Found {len(sheet_links)} product links in {sheet_name}")
                
                for i, link in enumerate(sheet_links):
                    product_code = link.url.split('/')[-1] if '/' in link.url else link.url
                    
                    row_data = {
                        'Product Code': product_code,
                        'EAN Code': '',
                        'RAL Number': '', 
                        'Net Width (mm)': '',
                        'Net Height (mm)': '',
                        'Net Depth (mm)': '',
                        'Package Units': '',
                        'Package Weight (kg)': ''
                    }
                    
                    # Fill in extracted data if available
                    if link.extracted_data:
                        total_links_with_data += 1
                        data = link.extracted_data
                        row_data.update({
                            'EAN Code': data.ean or '',
                            'RAL Number': data.ral_number or '',
                            'Net Width (mm)': data.net_width or '',
                            'Net Height (mm)': data.net_height or '',
                            'Net Depth (mm)': data.net_depth or '',
                            'Package Units': data.package_units or '',
                            'Package Weight (kg)': data.package_weight or ''
                        })
                        self.logger.debug(f"✅ Product {product_code}: Has extracted data")
                    else:
                        total_links_without_data += 1
                        self.logger.debug(f"❌ Product {product_code}: No extracted data (web scraping needed)")
                    
                    result_data.append(row_data)
                
                # Create dataframe with only new columns
                if result_data:
                    enhanced_df = pd.DataFrame(result_data)
                else:
                    # If no links found, create empty dataframe with column structure
                    enhanced_df = pd.DataFrame(columns=new_columns)
                
                # Write the results-only dataframe to Excel
                enhanced_df.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.info(f"📄 Sheet {sheet_name}: {len(result_data)} rows written with ONLY new columns")
        
        self.logger.info(f"📊 SUMMARY:")
        self.logger.info(f"   📝 Results-only Excel file saved: {output_path}")
        self.logger.info(f"   ✅ Links with extracted data: {total_links_with_data}")
        self.logger.info(f"   ❌ Links without data (need web scraping): {total_links_without_data}")
        self.logger.info(f"   📋 CONTAINS ONLY NEW COLUMNS - NO ORIGINAL DATA")
        
        return output_path